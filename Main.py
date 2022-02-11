# -*- coding: utf-8 -*-



import os
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *
from PyQt5.uic import loadUi
import pandas as pd
import numpy as np
import Functions_Project as functions
from MCDA_Project_1 import Ui_MainWindow
import openpyxl
from PyQt5 import *


###############################################################################################################################################


 
class main(QMainWindow): #create the class that type of QMainWindow that is in the PyQt5 lib.
    
    array_XlsStr, array_XlsFlt, hHead, vHead, array_WCalculate = [], [], [], [], []
    #yöntem hesaplamalarında kullanacak olduğumuz arrayler olacak listeler
    boolean, weight, w_Find = False, False, False
    #fonksiyonlarda lazım olacak bool türü değişkenler
    fileName, fileWay, way, error_msg, err_msg = "", "", "", "", "NONE"
    #dosya işlemlerinde kullanacağımız dosya yolu isim ve yöntem isminin tutulacağı değişkenler
    callnumber, counter, select_row, select_column, w_condition = 1, 1, -1, -1, 0
    #excelde yazdırmak için ihtiyacımız olacak sayaclar
    return_result_1,  return_result_2, return_result_3, return_result_4, return_result_5, return_result_W = [], [], [], [], [], []
    callCount = 0
    #sonuçları excel e yazdırmak için kullanılacak diziler

###############################################################################################################################################

    def __init__(self):#Create the init func. Firstly this func. runs.
        super().__init__() #Calling the parent class' init func. Secondly this func. runs.

        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.ui.btn_hesapla.setEnabled(False)#hesapla butonu table a excelden veri yüklenmediği sürece çalışır durumda olmayacaktır
        self.ui.tabWidget.setTabEnabled(1, False)#ağırlık hesaplamak için olan tab ı erişilemez yaptık
        self.ui.tabWidget.setCurrentIndex(0)
        self.ui.btn_wCalculate.setEnabled(True)

###############################################################################################################################################

    def checked_CBox(self):#checkbox seçilme slotu
        ix = self.ui.tableWidget_xlsTable.rowCount()
        isThere = self.ui.tableWidget_xlsTable.verticalHeaderItem(ix-1).text()
        
        if self.ui.cBox_isThereWeight.isChecked() and isThere != "W" and isThere != "WEIGHT" and isThere != "weight" and isThere != "Weight":
            #ischecked metodu geriye bool döndürür
            self.ui.btn_weightCalculete.setEnabled(True)
            #checkbox işaretli ise ağırlık hesaplama butonu aktif hale gelecektir
            #tersi halinde aktif edilmeyecektir.
        else:
            self.warning_message("GEÇERLİ TABLODA AĞIRLIK DEĞERİ BULUNMAKTADIR")
            self.ui.btn_weightCalculete.setEnabled(False)

###############################################################################################################################################
        
    def btn_fileway_clicked(self):#created by signal-slot way in qtDesigner
        if(self.ui.tableWidget_xlsTable.verticalHeaderItem(1) != None or self.ui.tableWidget_results.verticalHeaderItem(1) != None):
            #eğer table larda veri varsa table ların için her dosya seçiminde temizlenecektir.
            self.ui.tableWidget_xlsTable.clear()
            self.ui.tableWidget_results.clear()
            
        options = QFileDialog.Options()#qfiledialog objesi oluşturulur
        #getopenfilename metodu dosya açmak için kullanılan bir filedialog tur.
        files = QFileDialog.getOpenFileName(self, "Dosya Aç" , #açılan pencerenin başlığı
                                            os.getenv("HOME") , #varsayılan olarak açılacağı konum
                                            "Excel Dosyaları (*.xls , *.xlsx)",#seçilebilecek dosya türü
                                            options=options )#this method return tuple type data
        selected_file_name = files[0]#the data in the index 0 of files has copied to the str type parameter
        start = selected_file_name.rfind("/")#en / işeretinin olduğu index değeri 
        end = selected_file_name.rfind(".")#en son . işeretinin olduğu index değeri
        self.fileName = selected_file_name[start+1:end]+"_sonuc.xlsx"#dosya kaydederken kullanılacak default dosya ismi için str birleştirme yapıyoruz
        #tutulan start ve end deki indexlerin arasında açılan dosyanın uzantısız ismi bulunmaktadır.
        self.ui.label_selectedFileName.setText(selected_file_name)#set the text that is the file way and name on the label
        self.loadExcelData(selected_file_name)#call the func that makes the filling the qtablewidget

###############################################################################################################################################
        
    def find_items(self, value_zero):#tablo içinde aranan değerin olup olmadığı kontrol edilir.
    #verilen değerin tamamiyle aynı olması gerektiğine dair bir fonksiyon yazıldı
  
        items = self.ui.tableWidget_weights.findItems(value_zero, QtCore.Qt.MatchExactly)#bu metod dizi geri döndürür
        brush = QtGui.QBrush(QColor("red"))#arama sonucu bulunan değerler kırmızı renkli olarak değiştirilir
        brush.setStyle(Qt.SolidPattern)
        for item in items:
            item.setForeground(brush)
        if len(items)==0:#aranan veeeriden bulunmadıysa false döner
            return False
        else:
            return True

###############################################################################################################################################
        
    def btn_save_results_clicked (self):
       self.saveFileDialog()#dosya kaydetme fonksiyonu çağrılır.
       
       self.callnumber = 1#metodu çağırma sayımızın tutulacağı sayıcın değer ataması yapılmaktadır.
       #her dosya kaydetme butonuna basıldığında sayac 1 den başlayacaktır.
       #kaydedilecek her array için excele veri kaydetme fonksiyonu çağrılır.
       if self.weight == True and self.w_Find ==True:
              self.exportExcel(self.return_result_W)
           
       else:
             
              self.exportExcel(self.return_result_1)
              self.exportExcel(self.return_result_2)
              self.exportExcel(self.return_result_3)
              self.exportExcel(self.return_result_4)
              #bütün yöntemler 4 adet array e sahiptir o yüzden 4 kez fonksiyon 4 farklı array için çağrılır.
              if self.way == "TOPSIS" or self.way == "VIKOR":
                  #eğer seçilen yöntem topsis ya da vikor ise 1 kere daha fonksiyon 5. array için çağrılır
                  self.exportExcel(self.return_result_5)

###############################################################################################################################################

    def loadExcelData(self, excel_file_dir):#this func takes a file way that is the str type
        vHeader, hHeader = [], []
        self.df = pd.read_excel(excel_file_dir)# makes a dataframe (with the excel data) with pandas lib
        if self.df.size == 0:# is the table empty
            return
        
      #vertical header ı tutmak için döngü oluşturulur
        i=0  
        for row in self.df.iterrows():
            values = row[1]#row un 1. indexinde satırın değerleri tutulmaktadır
            if(i<self.df.shape[0]):# Gives number of rows
            #eğer döngü sayısı dataframe in satır sayısından küçükse değer eklemesi yapılır vHead listesine
                self.vHead.append(str(values[0]))#satırın ilk indexindeki değer 0. sütundaki satır değeridir.
                i+=1#sayaç 1 arttırılır
            else:#döngü sayısı satır sayısına eşit olduğunda döngüden çıkılır.
                break
            
        self.df.fillna('', inplace=True) # if a cell is empty the empty cell fill with the ''
        self.ui.tableWidget_xlsTable.setRowCount(self.df.shape[0])#row count
        #tablewidget için satır ve sütun sayısı belirtilir
        self.ui.tableWidget_xlsTable.setColumnCount(self.df.shape[1]-1)#column count
        #print(self.df.columns)#başlık satırını döndürür
        self.ui.tableWidget_xlsTable.setHorizontalHeaderLabels(self.df.columns[1:])#ilk satır başlık olur
        self.hHead = np.array(self.df.columns, dtype=object)#horizontal header ı global değişkene atıyoruz
        self.ui.tableWidget_xlsTable.setVerticalHeaderLabels(self.vHead)
        #yukarıda bulunan alternatifler sütunu(vertical header) vertical header olarak tabloya eklenir.
      
     
        # returns to pandas array object
        for row in self.df.iterrows():
            values = row[1]

            for col_index, value in enumerate(values):# With the enumerate method, makes a tuple containing index and values of excel data
                if(col_index != 0):#0. sütunda vertical header değerleri bulunmaktadır o sütunu es geçiyoruz.
               
                #if isinstance(value, (float)):#is there a int or float type data
                      #  value = '{0:0,.0f}'.format(value) #if "if" method is true, data will be convert to int or float
                    # print(row[0]) 
                    tableItem = QTableWidgetItem(str(value)) # if the data is not a number, tha data convert to str
                    self.ui.tableWidget_xlsTable.setItem(row[0], col_index-1, tableItem)# in the row[0] takes row indexs of the excel table data
                    self.ui.tableWidget_xlsTable.resizeColumnsToContents() #the method does the resize the table size
        
        self.fillArray_withTData()#table a eklenen değerler global değişkenler yardımı ile array e eklenir.
        # burada değerler yatay ve dikey başlıklar olmadan array e eklenir. Yani sayısal veriler array e alınır

###############################################################################################################################################

    def fillArray_withTData(self):#excelden gelen verilerin yüklendiği table dan verileri arraya ekliyoruz.
    #ben table ın hücrelerinde değişiklik yapmaya izin verdim o yüzden table dan array e veri çekiyoruz
        if self.weight == True:
            rowCount = self.ui.tableWidget_weights.rowCount()
            columnCount = self.ui.tableWidget_weights.columnCount()
        else:
            
            rowCount =int( self.ui.tableWidget_xlsTable.rowCount())
            #table ın satır ve sütun sayılarını alıyoruz.
            columnCount=int(self.ui.tableWidget_xlsTable.columnCount())
            
        array_xls_str=np.empty((rowCount,columnCount), object)
        #alınan satır ve sütun sayısı değerleriyle kendimize yeni iki adet boş object türünde array tanımlıyoruz.
        #eklenen verinin türü aynı kalması ve dinamik olabilmesi çin object türü kullanıyoruz.
        array_xls_flt=np.empty((rowCount-1,columnCount), object)
        # burada 1 satır eksik boyut belirliyoruz çünkü ağırlık değerlerinin olduğu son satırı istemiyoruz.
        
        for row in range(rowCount):
            for column in range(columnCount):
                if self.weight ==True:
                    item=self.ui.tableWidget_weights.item(row,column).text() 
                else:
                    item=self.ui.tableWidget_xlsTable.item(row,column).text()  #table da belirtilen indexte bulunan değeri değişkene atıyoruz           
                
                array_xls_str[row, column] = item#arrayde aynı index e değişkende tutulan str türündeki değeri atıyoruz
        if self.weight != True:
            self.array_XlsStr, self.err_msg = functions.convert_to_numerical(self, array_xls_str, columnCount, rowCount)#dönüştürülen verileri globaldeki arraye aktarıyoruz
            #arraydeki str türü sayısal verileri float a dönüştürmek için functions modülünden dönüştürme fonksiyonunu çağırıyoruz.
            self.ui.btn_hesapla.setEnabled(True)#arraye veri ekleme tamamlandığında hesapla butonunu çalışır hale getiriyoruz
            self.array_XlsFlt = np.copy(array_xls_flt)#oluşturulan boş array in kopyasını global daki tanımlanmış değişkene atıyoruz
        else:
            self.array_WCalculate, self.err_msg = functions.convert_to_numerical(self, array_xls_str, columnCount, rowCount)
            # self.array_WCalculate = np.copy(array_xls_flt)

###############################################################################################################################################
    
    def fillTable_wArrayData(self, Head1, array):
        #en son aşamadaki yöntem sonucunda 
        # bir array ve onun alternatiflerin olduğu vertical header ı bulunur onun için o ikisini parametre olarak istiyoruz
        if self.weight == False and self.w_Find == False:#bu fonksiyon eğer ağırlık hesaplamak için olan table için kullanılmayacaksa if içine girecek
            self.ui.tableWidget_results.setRowCount(Head1.shape[0])#başlık satırı kadar satır sayısı atanır table a 
            self.ui.tableWidget_results.setColumnCount(array.shape[1])#array in sütun sayısı kadar sütun sayısı atanır table a 
            self.ui.tableWidget_results.setVerticalHeaderLabels(Head1[:])#alınan header ı vertical header olarak table a atıyoruz
        elif self.weight == True and self.w_Find == False :
            self.callCount = array.shape[1]
            self.ui.tableWidget_weights.setRowCount(array.shape[1])
            self.ui.tableWidget_weights.setColumnCount(array.shape[1])
            self.ui.tableWidget_weights.setVerticalHeaderLabels(Head1[:])
            self.ui.tableWidget_weights.setHorizontalHeaderLabels(Head1[:])
            
        elif self.weight == True and self.w_Find == True:
            # tableWidget_resultWeight.setColumnCount(2)
            self.ui.tableWidget_resultWeight.setRowCount(array.shape[0])
            # self.ui.tableWidget_weights.setVerticalHeaderLabels(Head1[:])
            
        
            
        for j in range(array.shape[1]):#sütun sayısı
            for i in range(Head1.shape[0]):#satır sayısı
            
                tableItem = QTableWidgetItem(str(array[i,j]))
                #table a veriler str türünde atanır onun için değişkene tableverisi oluşturuyoruz atamak için
               
                if self.weight == True and self.w_Find == False:
                    self.ui.tableWidget_weights.setItem(i, j, tableItem)
                    self.ui.tableWidget_results.resizeColumnsToContents()
                    
                elif self.weight == False and self.w_Find == False:
                    # if i==j:
                    #     tableItem.setFlags(tableItem.flags() ^ Qt.NoItemFlags)
                    #     self.ui.tableWidget_results.setItem(i, j, tableItem)#verilen index cell ine tableitem atanır
                    # else:
                    self.ui.tableWidget_results.setItem(i, j, tableItem)
                    self.ui.tableWidget_results.resizeColumnsToContents()
                    
                elif self.weight == True and self.w_Find == True:
                    self.ui.tableWidget_resultWeight.setItem(i, j, tableItem)
                    self.ui.tableWidget_results.resizeColumnsToContents()
                    self.w_Find == False
                        #sütun boyutunun mikatrına göre dinamik değişmesine olanak veren bir method çağırıyoruz
        # self.weight = False #tabloya veri yükleme işlemi bittikten sonra değişkenin false olarak atamasını yapıyoruz.
        return True
    
    # def getvHead(self):
    #     self.vHead = []
    #     self.ui.tableWidget_xlsTable.rowCount()
    #     for i in range(self.ui.tableWidget_xlsTable.rowCount()):
    #         self.vHead.append(self.ui.tableWidget_xlsTable.verticalHeaderItem(i).text())

###############################################################################################################################################

    def cellValue_WCalculate (self):
      if 15 >= self.ui.tableWidget_xlsTable.columnCount():
          
          self.ui.tabWidget.setTabEnabled(1, True)#ağırlık hesaplamak için olan tab ı erişilir yaptık
          #eğer ağırlık hesaplama butonuna tıklandıysa o sırada tablewidgetta yüklü olan tabloya göre bir matris oluşturulur
          self.weight = True #tablo doldurmak için kullanılan fonksiyonda hangi tablo için işlem yapılacağını bu değişken ile belirliyoruz
          #default olarak false ayarlanmıştır. ağırlık hesaplama butonuna tıklandığı taktirde true olarak değiştirilir
          arr_ForWeight = np.eye(self.hHead.shape[0]-1)#kriter miktarına göre boyut veriliyor
          self.fillTable_wArrayData(self.hHead[1:], arr_ForWeight)
          
          self.ui.tabWidget.setCurrentIndex(1)
      else:
          self.warning_message("EN FAZLA 15 ADET KRİTERİN AĞIRLIĞI HESAPLANABİLMEKTEDİR.")

###############################################################################################################################################
    def is_int (self, value):
        try:
            int(value)
            return True#değer dönüştülebiliyorsa true değer döndürür
        except ValueError:
            return False

###############################################################################################################################################

    def findWeight(self):#tabloda düzenlenmiş matrisin ağırlık hesaplama butonu için slotu
        try:
            if self.err_msg != "NONE":
                self.warning_message(err_msg)
                self.err_msg = ""
            else:
                
                isThere_01 = self.find_items("0")
                isThere_02 = self.find_items("0.0")
                if isThere_01 == False and isThere_02 == False:#tabloda 0 değeri olmamalı onun için kontrol ediyoruz
                    self.fillArray_withTData()#tablodan array a veri ekleme fonksiyonu
                    self.return_result_W = np.array(self.return_result_W, object)
                    self.return_result_W, self.w_condition, self.err_msg = functions.weight_Calculate(self, self.array_WCalculate, self.hHead[1:])
                    self.ui.label_weightRank.setText(str(self.w_condition))
                    if self.err_msg != "NONE":
                        #Default olarak değişkene none verisi atandı eğer hata oluşursa içerik değişimini kontrol ediyoruz bu şekilde
                        self.warning_message(err_msg)

                    else:
                        #hatasız geldiyse bu satıra kadar fonksiyon w_find değişkenine true değeri atanır 
                        #ve hesaplama yapılabildiğine dair işaret vermiş oluyoruz
                        self.w_Find = True
                        self.fillTable_wArrayData(self.hHead[1:], self.return_result_W)
                        self.ui.btn_saveWeights.setEnabled(True)#dosya kaydetme butonunu erişilebilir kılıyoruz
                else:
                    self.warning_message("TABLODA HALA 0 DEĞERİ BULUNMAKTADIR. LÜTFEN BİR DEĞER GİRİNİZ.")
        except Exception:
            self.warning_message("AĞIRLIK MATRİSİ İSTENEN FORMATTA DEĞİLDİR. DÜZENLEYİP TEKRAR DENEYİNİZ.")

###############################################################################################################################################
    def cellValue(self, new_item):#doubleclicked sinyali için çalışır ağırlık hesabı yapılan tablewidget ta 

        self.select_row = new_item.row()
        self.select_column = new_item.column()
        # print(self.select_row, self.select_column)

###############################################################################################################################################
    def cellValue_change(self, cell_value):
       
        row = cell_value.row()
        column = cell_value.column()
        # item_1 = self.ui.tableWidget_weights.item(row, column).text()
        # print("****", item_1) 
       
        
        if self.is_int(cell_value.text())==True and float(cell_value.text()) != 0.0 and row != column:
            if self.select_row==row and self.select_column == column:
                value = 1 / float(cell_value.text())
                tableItem = QTableWidgetItem(str(value))
                self.ui.tableWidget_weights.setItem(column, row, tableItem)
                # QMessageBox.Warning(self, "Kritik Mesaj", "Girilen Değer Sayısal Veri Değildir. Yeniden Giriniz.", QMessageBox.Ok)
                
        elif row != self.select_row and self.select_column != column and functions.is_number(cell_value.text()) == True :
            pass
        
        # else:
        #     self.error_msg = "Girilen Değer Pozitif Tam Sayı Değildir. Yeniden Giriniz."
        #     self.warning_message(self.error_msg)

###############################################################################################################################################
    def current_item(self, new_item, old_item):
        #seçilen hücre sinyali için slot yeni aktive olan ve en son deaktive olan değerleri gönderir
        # itemların indexleri alınır
        n_row = new_item.row()
        n_col = new_item.column()
        o_row = old_item.row()
        o_col = old_item.column()
        doubled_item = self.ui.tableWidget_weights.item(self.select_row, self.select_column).text()
        #çift tıklama sinyalinde alınan index değerleri 
        # print("new", new_item.text(), "old", old_item.text())
        if  n_row == n_col:
            if self.ui.tableWidget_weights.item(n_row, n_col).text() != "1.0":
                tableItem = QTableWidgetItem("1.0")
                self.ui.tableWidget_weights.setItem(n_row, n_col, tableItem)
               
        if o_row == o_col:
            if self.ui.tableWidget_weights.item(o_row, o_col).text() != "1.0":
                tableItem = QTableWidgetItem("1.0")
                self.ui.tableWidget_weights.setItem(o_row, o_col, tableItem)
        
        if self.is_int(doubled_item)==False and self.select_row != self.select_column:
            #çift tıklanarak girilen değerin int olup olmadığı ve indexlerin birbirine eşit olup olmadığına bakılır
            #girilen değer int değilse 0 değeri atanır hücreye 
            tableItem = QTableWidgetItem(str(0))
            self.ui.tableWidget_weights.setItem(self.select_row, self.select_column, tableItem)
            self.error_msg = "Girilen Değer Pozitif Tam Sayı Değildir. Yeniden Giriniz."
            self.warning_message(self.error_msg)#hata mesaj penceresi eklenir 
       
        # if self.error_msg.count("Pozitif")==1 and n_row == self.select_row and self.select_column == n_col and self.select_column != self.select_row:
        #     if self.is_int(self.ui.tableWidget_weights.item(n_row, n_col).text())==True:
        #         tableItem = QTableWidgetItem(str(0.0))
        #         self.ui.tableWidget_weights.setItem(n_row, n_col, tableItem)
        # elif self.error_msg.count("Pozitif")==1 and o_row == self.select_row and self.select_column == o_col and self.select_column != self.select_row:
        #     if self.is_int(self.ui.tableWidget_weights.item(o_row, o_col).text())==True:    
        #         tableItem = QTableWidgetItem(str(0))
        #         self.ui.tableWidget_weights.setItem(o_row, o_col, tableItem)
###############################################################################################################################################
  
    def calculate(self):#yöntem hesaplamak için olan hesapla butonunun slotu
        
        try:
            end_row = self.ui.tableWidget_xlsTable.rowCount()
            isThere1 = self.ui.tableWidget_xlsTable.verticalHeaderItem(end_row-1).text()
            if isThere1 != "W"and isThere1 != "WEIGHT" and isThere1 != "weight" and isThere1 != "Weight" :#en son satırda ağırlık satırı kontrolü yapılıyor.
                self.warning_message("Ağırlık Değerleri Yok ya da Son Satırda Yer Almamaktadır.\nLütfen Düzenleyip Yeniden Yükleyiniz.")
                #eğer ağırlık yoksa uyarı mesajı gelecektir ekrana
            
            # if self.vHead[-1]!="W" and self.vHead[-1]!="weight" and self.vHead[-1]!="Weight" and self.vHead[-1]!="WEIGHT":
            else:
                self.weight = False
                self.w_Find = False
                isThere1 = self.find_items("0")
                isThere2 = self.find_items("0.0")
                if isThere1 == False and isThere2 == False:
                
                    self.fillArray_withTData()#table a eklenen değerler global değişkenler yardımı ile array e eklenir.
                    # burada değerler yatay ve dikey başlıklar olmadan array e eklenir. Yani sayısal veriler array e alınır
                    outcome = np.sum(self.array_XlsStr[-1,:])#ağırlık değerleri toplamını buluyoruz
                    print(outcome)
                    if 0.9 <= outcome < 1.1:#ağırlık değerleri toplamı 1 e eşit olmalıdır. ancak tam 1 çıkmadığı ama yakın olduğu durumlar olabilir.
                        self.ui.tableWidget_results.clear()#sonuçların tutulduğu tablewidget her hesapla butonu tıklandığında temizlenmesi sağlanır
                        #global de oluşturulan diziler array e dönüştürülür.
                        self.return_result_1 = np.array(self.return_result_1, object)
                        self.return_result_2 = np.array(self.return_result_2, object)
                        self.return_result_3 = np.array(self.return_result_3, object)
                        self.return_result_4 = np.array(self.return_result_4, object)
                        self.return_result_5 = np.array(self.return_result_5, object)
    
                        # print( self.array_XlsStr, self.array_XlsFlt, self.hHead, self.vHead)
                        #combobox taki en son seçili olan değerin indexine göre konrol edilip yöntemler uygulanır
                        if(self.ui.cBox_yontem.currentIndex() == 0):#TOPSIS
                            #geri döndürülen değerlerin anlaşılır değişken isimleriyle ataması sağlanır.
                            rankArray, norm_array, weight_array, siPlus, siMinus = functions.TOPSIS(self,  self.array_XlsStr, self.array_XlsFlt, self.hHead, self.vHead)
                            self.boolean = self.fillTable_wArrayData( rankArray[:,0], rankArray[:,1:])
                            #en son sonuç değerlerinin tutulduğu vertical header ve matris sonuç table ına yüklenir
                            #globalde tanımlanan değişkenlere geri döndürülen değerlerin kopyaları atanır.
                            #aslında hafızada fazladan alan tuttuklarının farkındayım ancak kodlara tekrar bakıldığında
                            # matrislerin hangi aşamalara sahip matrisler olduğunu anlamak daha kolay o yüzden bu şekilde yaptım
                            #eğer proje büyütülürse bu kopya ataması yerine direk global değişkenlere atama yapılabilir.
                            self.return_result_1 = np.copy(norm_array)
                            self.return_result_2 = np.copy(weight_array)
                            self.return_result_3 = np.copy(siPlus)
                            self.return_result_4 = np.copy(siMinus)
                            self.return_result_5 = np.copy(rankArray)
                            self.way = "TOPSIS"#globalde en son işlem yapılan yöntem adı tutulur farklı bir fonksiyonda koşul olarak kontrolü sağlanacaktır.
                        elif(self.ui.cBox_yontem.currentIndex() == 1):#ARAS
                            ranked, array, norm_arr, W_array = functions.ARAS(self,  self.array_XlsStr, self.array_XlsFlt, self.hHead, self.vHead)
                            self.boolean = self.fillTable_wArrayData(ranked[1:,0], ranked[1:,1:])
                            self.return_result_1 = np.copy(array)
                            self.return_result_2 = np.copy(norm_arr)
                            self.return_result_3 = np.copy(W_array)
                            self.return_result_4 = np.copy(ranked)
                            self.way = "ARAS"
                        elif(self.ui.cBox_yontem.currentIndex() == 2):#VIKOR
                            array, linear, array_weight, result_si_ri, qi_outcome_sort = functions.VIKOR(self,  self.array_XlsStr, self.array_XlsFlt, self.hHead, self.vHead)
                            self.boolean = self.fillTable_wArrayData( qi_outcome_sort[:,0], qi_outcome_sort[:,1:])
                            self.return_result_1 = np.copy(array)
                            self.return_result_2 = np.copy(linear)
                            self.return_result_3 = np.copy(array_weight)
                            self.return_result_4 = np.copy(result_si_ri)
                            self.return_result_5 = np.copy(qi_outcome_sort)
                            self.way = "VIKOR"
    
                        self.ui.btn_writeResults.setEnabled(self.boolean)#sonuç yazdır butonunu enable haline getiriyoruz.
                        #tabi fonksiyon bu satıra kadar çalışması sağlanırsa
                        self.boolean = False
                        self.fillArray_withTData()
                        #burada global değişkenlerle array lere değer atandığı için en son yapılan işlemler globaldeki arrayleri etkilemektedir
                        # bu yüzden her hesaplama yapılacak olduğunda baştan tabledan veriler array a çekilmektedir.
                    else:
                        self.warning_message("AĞIRLIK DEĞERLERİ TOPLAMININ 1 DEĞERİNE UZAKLIĞI 0.1'DEN FAZLADIR. AĞIRLIK DEĞERLERİNİ DÜZENLEYİNİZ")
                else:
                    self.warning_message("TABLODA HALA 0 DEĞERİ BULUNMAKTADIR. LÜTFEN BİR DEĞER GİRİNİZ.")
        except Exception:
            self.warning_message("MATRIS OLMASI BEKLENEN FORMATTA DEĞİLDİR. VERİLERİNİZİ TEKRAR KONTROL EDİNİZ")

###############################################################################################################################################

    def addRowInTable(self):#ağırık satırı ekleme foksiyonu
        self.ui.tabWidget.setCurrentIndex(0)#tabwidget ın 0. indexini ekrana getirir
        
        # currentRowCount = self.ui.tableWidget_xlsTable.rowCount()#mevcut satır sayısını alırız
        end_row = self.ui.tableWidget_xlsTable.rowCount()
        isThere1 = self.ui.tableWidget_xlsTable.verticalHeaderItem(end_row-1).text()
        if isThere1 != "W" and isThere1 != "WEIGHT" and isThere1 != "weight" and isThere1 != "Weight" :#en son satırda ağırlık satırı kontrolü yapılıyor.
            self.ui.tableWidget_xlsTable.insertRow(end_row)#yeni satırı sona eklemek için üst satırda alınan satır sayısını kullanıyoruz
            self.ui.tableWidget_xlsTable.setVerticalHeaderItem(currentRowCount, QTableWidgetItem(str("W")))
            self.vHead = np.append(self.vHead, ["W"], axis = 0)
            # eklenen satırın başlığını "W" olarak atıyoruz
        
        for i in range (self.ui.tableWidget_xlsTable.columnCount()):#sütun sayısı kadar döngü oluşturulur
            self.ui.tableWidget_xlsTable.setItem(currentRowCount, i, QTableWidgetItem(str(self.return_result_W[i,1])))
            # son satırın her bir sütun indexine ağırlık değerlerinin tutulduğu arraydan veri eklenir 

###############################################################################################################################################

    def exportExcel(self, array):
        # book = Workbook()
        wb = openpyxl.load_workbook(filename=self.fileWay)#kaydedilmiş bir xls dosyasını açmak için method kullanıyoruz
        #verilen dosya yolu ile dosyayı açıyor bu sayede 
        sheet = wb.active#yazılacak excel sayfasını aktif hale getiriyoruz.
        if self.weight != True and self.w_Find != True:
            #tutulan callnumber sayacı ile fonksiyonun çağrılma sayısını 1 den başlayarak tutuyoruz
            #sayac bize eklenecek matrisin hangisi olduğu hakkında ipucu olacaktır. 
            #yöntem ismi ve çağrılma sayısı kontrolleri ile hangi matrisin kayıt edileceği anlaşılır bu sayede bir başlık eklenir
            if (self.callnumber == 1 and self.way == "TOPSIS") or (self.callnumber == 2 and self.way!="TOPSIS" ):
                #sheet.merge_cells hücre birleştirme komutudur sütun satısı kadar cell birleştirilir başlık metni eklenir sonrasında
                #excelde veri yazma indexi 1 den başlamaktadır.
                sheet.merge_cells(start_row=self.counter, start_column=1, end_row=self.counter, end_column=array.shape[1]+1)
                sheet.cell(row=self.counter, column=1, value="NORMALLEŞTİRİLMİŞ MATRİS")
                #sheet.cell bize belli bir hücreye veri eklenmesinde yardımcı olacaktır.
                self.counter+=1
            elif (self.callnumber == 2 and self.way == "TOPSIS") or (self.callnumber == 3 and self.way != "TOPSIS"):
               sheet.merge_cells(start_row=self.counter, start_column=1, end_row=self.counter, end_column=array.shape[1]+1)
               sheet.cell(row=self.counter, column=1, value="AĞIRLIKLANDIRILMIŞ MATRİS")
               self.counter+=1
            elif (self.callnumber == 5 and self.way != "ARAS") or (self.callnumber == 4 and self.way == "ARAS"):
                sheet.merge_cells(start_row=self.counter, start_column=1, end_row=self.counter, end_column=array.shape[1]+2)
                #buradaki başlık uzun olduğu için 1 hücre fazla eklenerek birleştirilme yapılmıştır
                sheet.cell(row=self.counter, column=1, value="KÜÇÜKTEN BÜYÜĞE SIRASI BELİRTİLMİŞ MATRİS")
                self.counter+=1
            elif(self.callnumber == 3 and self.way == "TOPSIS"):
                sheet.merge_cells(start_row=self.counter, start_column=1, end_row=self.counter, end_column=array.shape[1]+1)
                sheet.cell(row=self.counter, column=1, value="Si PLUS DEĞERLERİ")
                self.counter+=1
            elif(self.callnumber == 4 and self.way == "TOPSIS"):
                sheet.merge_cells(start_row=self.counter, start_column=1, end_row=self.counter, end_column=array.shape[1]+1)
                sheet.cell(row=self.counter, column=1, value="Si MINUS DEĞERLERİ")
                self.counter+=1
                
        r = 0#array için satır indexi tanımlıyoruz
        for row in range(self.counter, self.counter+array.shape[0]):#
            for column in range(array.shape[1]):
                sheet.cell(row=row, column=column+1).value = array[r,column]
            r+=1
        wb.save(self.fileWay)#açılan dosya kaydediliyor
        wb.close()#açılan dosya kapatılıyor
        
        #en son kalınan satır sayısının 1 fazlası yeniden yazma işleminde başlangıç satırı olacağı için sayacı bu şekilde arttırıyoruz.
        #bu sayac ile eklenecek matrislerin aşamadaki başlıklarını eklemek için bize sıra belirtecektir.
        if self.w_Find == True:
            self.w_Find = False
            
        else:
            self.counter+=array.shape[0]+1
            self.callnumber+=1

###############################################################################################################################################
            
    def saveFileDialog(self):
        options = QFileDialog.Options()#qfiledialog objesi oluşturulur.
        files, _ = QFileDialog.getSaveFileName(self,"EXCEL DOSYASI KAYDET", self.fileName, "Excel Dosyaları (*.xls , *.xlsx)", options=options)
        #getsavefilename metodu bize dosya kaydetmekte yardımcı olacaktır. 
        #ilk değer filedialogta görünen pencerenin başlığıdır. filename varsayılan dosya ismidir
        #üçüncü değer kaydetme için belirtilen dosya türüdür.
        #en sona da objemizi ekliyoruz.
        self.fileWay = files
        # Kaydedilen dosya yolu tutulur
        
        #qfiledialog ile dosyayı kaydedemediğim için 
        # kaydederken kullanılan dosya yolunu kullanarak openpyxl kütüphanesi ile dosyayı kaydettim
        #creates a new workbook
        wb = openpyxl.Workbook()
        #saving the workbook
        wb.save(self.fileWay)
        # print(self.fileWay, _)

###############################################################################################################################################
   
    def warning_message(self, error):#uyarı mesajları için messagebox
        # QMessageBox.Warning(self, "Kritik Mesaj", "Girilen Değer Sayısal Veri Değildir. Yeniden Giriniz.", QMessageBox.Ok)
        QMessageBox.warning(self, "UYARI", error, QMessageBox.Ok)


###############################################################################################################################################
    
if __name__ =="__main__":
    app = QApplication([])
    window = main()
    window.show()
    app.exec_()  






